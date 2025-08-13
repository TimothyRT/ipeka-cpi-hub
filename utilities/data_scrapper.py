import pandas as pd
from openpyxl import *
from openpyxl.cell.cell import Cell
from sqlalchemy import create_engine

from datetime import datetime


SPECIAL_POSITIONS = ["ADMINISTRATION", "CHAPLAIN", "COUNSELOR", "FINANCE", "F&P", "ICT", "IT SUPPORT", "KEPALA TU", "LIBRARY", "STORE"]


def format_gender(inp: str):
    return inp[0].upper()


def format_position(inp: str):
    inp = inp.strip().upper()
    if "TEACHER" in inp:
        return "TEACHER"
    if "PRINCIPAL" in inp:
        if "VICE" in inp:
            return "VICE PRINCIPAL"
        else:
            return "PRINCIPAL"
    if inp in SPECIAL_POSITIONS:
        return inp
    return "TEACHER"


def formatify(inp: (str | int | datetime | None)):
    match str(type(inp)):
        case "<class 'int'>":
            return str(inp)
        case "<class 'str'>":
            return inp
        case "<class 'datetime.datetime'>":
            return inp.strftime("%Y-%m-%d")
        case "<class 'NoneType'>":
            return ""
        case _:
            raise Exception(str(type(inp)))


def scrap_employee_xlsx(input_file_path: str, sheet_name: str):
    wb = load_workbook(input_file_path, data_only=True)
    ws = wb[sheet_name]
    
    number_cell: (Cell | None) = None
    i = 0
    while number_cell is None:
        potential_number_cell = ws["B1"].offset(i, 0)
        if potential_number_cell.value == "Number":
            number_cell = potential_number_cell
        else:
            i += 1
    
    headers: dict[str: Cell] = {}
    i = 0
    while True:
        i += 1
        header_cell: Cell = ws[number_cell.coordinate].offset(0, i)
        if header_cell.value:
            headers[header_cell.value] = header_cell
        else:
            break
    
    df_dict = {
        "name": [],
        "nip": [],
        "grade": [],
        "position": [],
        "join_date": [],
        "date_of_birth": [],
        "gender": [],
        "phone": [],
        "email": []
    }
    
    i = 0
    while True:
        i += 1
        if headers["Name"].offset(i, 0).value:
            df_dict["nip"].append(formatify(headers["NIP"].offset(i, 0).value).strip())
            df_dict["grade"].append(formatify(headers["Grade"].offset(i, 0).value).strip().upper())
            df_dict["join_date"].append(formatify(headers["Join Date"].offset(i, 0).value).strip())
            df_dict["position"].append(format_position(headers["Position"].offset(i, 0).value))
            df_dict["date_of_birth"].append(formatify(headers["Date of Birth"].offset(i, 0).value).strip())
            df_dict["gender"].append(format_gender(headers["Gender"].offset(i, 0).value))
            df_dict["phone"].append(formatify(headers["Phone"].offset(i, 0).value).strip())
            df_dict["email"].append(formatify(headers["Email"].offset(i, 0).value).strip().lower())
            df_dict["name"].append(formatify(headers["Name"].offset(i, 0).value).strip().title())
        else:
            break
        
    df = pd.DataFrame(df_dict)
    df.to_csv("static/gsheet/employee.csv")
            
    # return df["position"].unique()
    return df
    

if __name__ == "__main__":
    df = scrap_employee_xlsx(
        input_file_path="static/gsheet/employee.xlsx",
        sheet_name="Summary",
    )
    df.to_sql(
        "employee",
        con=create_engine('sqlite:///instance/database.db'),
        if_exists="replace",
        index=True
    )
