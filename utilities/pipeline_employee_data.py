import pandas as pd
from openpyxl import *
from openpyxl.cell.cell import Cell
from sqlalchemy import create_engine, text

from datetime import datetime
import os
import shutil

from utilities.database_uri import get_database_uri
from utilities.downloader import download_employee_gsheet


SPECIAL_POSITIONS = ("ADMINISTRATION", "CHAPLAIN", "COUNSELOR", "FINANCE", "F&P", "ICT", "IT SUPPORT", "KEPALA TU", "LIBRARY", "LOCATION COORDINATOR", "STORE")


def format_gender(inp: str):
    return inp[0].upper()


def format_position(inp: str):
    inp = inp.strip().upper()
    if "TEACHER" in inp or "ICT" in inp:
        return "TEACHER"
    if "PRINCIPAL" in inp:
        if "VICE" in inp:
            return "VICE PRINCIPAL"
        else:
            return "PRINCIPAL"
    if inp in SPECIAL_POSITIONS:
        return inp
    return "TEACHER"


def formatify(inp: (str | int | datetime | None)):
    match str(type(inp)):
        case "<class 'int'>":
            return str(inp)
        case "<class 'str'>":
            return inp
        case "<class 'datetime.datetime'>":
            return inp.strftime("%Y-%m-%d")
        case "<class 'NoneType'>":
            return ""
        case _:
            raise Exception(str(type(inp)))
        
        
def generate_placeholder_avatar(name: str) -> None:
    name_propercase = name.strip().title()
    name_uppercase = name_propercase.upper().replace(' ', '_')   
    if not os.path.exists(f"static/images/processed/employees/{name_uppercase}.png"):
        shutil.copy("static/images/icons/blank_avatar.png", f"static/images/processed/employees/{name_uppercase}.png")
        print(f"{name_propercase}'s profile pic hasn't been uploaded!")
    else:
        print(f"{name_propercase}'s profile pic already exists!")

        
def append_koorlok(df_dict: dict) -> None:
    for grade in ("KG", "EL", "JH", "SH"):
        df_dict["nip"].append(None)
        df_dict["grade"].append(grade.upper())
        df_dict["join_date"].append(None)
        df_dict["position"].append("LOCATION COORDINATOR")
        df_dict["date_of_birth"].append(None)
        df_dict["gender"].append("F")
        df_dict["phone"].append("")
        df_dict["email"].append("")
        df_dict["name"].append("Vivianti Kristanto Jacob")
        generate_placeholder_avatar(df_dict["name"][-1])


def scrap_employee_xlsx(input_file_path: str, sheet_name: str):
    wb = load_workbook(input_file_path, data_only=True)
    ws = wb[sheet_name]
    
    number_cell: (Cell | None) = None
    i = 0
    while number_cell is None:
        potential_number_cell = ws["B1"].offset(i, 0)
        if potential_number_cell.value == "Number":
            number_cell = potential_number_cell
        else:
            i += 1
    
    headers: dict[str: Cell] = {}
    i = 0
    while True:
        i += 1
        header_cell: Cell = ws[number_cell.coordinate].offset(0, i)
        if header_cell.value:
            headers[header_cell.value] = header_cell
        else:
            break
    
    df_dict = {
        "name": [],
        "nip": [],
        "grade": [],
        "position": [],
        "join_date": [],
        "date_of_birth": [],
        "gender": [],
        "phone": [],
        "email": []
    }
    
    i = 0
    while True:
        i += 1
        if headers["Name"].offset(i, 0).value:
            df_dict["nip"].append(formatify(headers["NIP"].offset(i, 0).value).strip())
            df_dict["grade"].append(formatify(headers["Grade"].offset(i, 0).value).strip().upper())
            df_dict["join_date"].append(formatify(headers["Join Date"].offset(i, 0).value).strip())
            df_dict["position"].append(format_position(headers["Position"].offset(i, 0).value))
            df_dict["date_of_birth"].append(formatify(headers["Date of Birth"].offset(i, 0).value).strip())
            df_dict["gender"].append(format_gender(headers["Gender"].offset(i, 0).value))
            df_dict["phone"].append(formatify(headers["Phone"].offset(i, 0).value).strip())
            df_dict["email"].append(formatify(headers["Email"].offset(i, 0).value).strip().lower())
            df_dict["name"].append(formatify(headers["Name"].offset(i, 0).value).strip().title())
            generate_placeholder_avatar(df_dict["name"][-1])
        else:
            break
        
    append_koorlok(df_dict)
        
    df = pd.DataFrame(df_dict)
    
    order_map_position = {
        "LOCATION COORDINATOR": 0,
        "PRINCIPAL": 1,
        "VICE PRINCIPAL": 2,
        "TEACHER": 3
    }
    order_map_grade = {
        "KG": 0,
        "EL": 1,
        "JH": 2,
        "SH": 3,
        "RO": 4
    }
    df.sort_values(
        by=["grade", "position"],
        inplace=True,
        key=lambda col: (
            col.map(order_map_position).fillna(4)  # fill others with a high number
            if col.name == "position" else col.map(order_map_grade)
        )
    )
    
    df.reset_index(drop=True, inplace=True)
    df.index.name = "index"
    df.to_csv("static/gsheet/employee.csv")
        
    print('scrap_employee_xlsx() SUCCESS')
    return df


def setup_employee_database(df: pd.DataFrame):
    engine = create_engine(get_database_uri())
    with engine.connect() as conn:
        table_exists = engine.dialect.has_table(conn, "employee")
        if table_exists:
            row_count = conn.execute(text("SELECT COUNT(*) FROM employee")).scalar()
            if row_count > 0:
                print("Table 'employee' already has data! Aborting.")
                return

        # If table hasn't existed, or table has zero rows
        success_rows = df.to_sql("employee", con=engine, if_exists="append", index=False)
        if success_rows and success_rows > 0:
            print('setup_employee_database() SUCCESS')


def run_employee_data_pipeline() -> None:
    download_employee_gsheet(
        link="https://docs.google.com/spreadsheets/d/1EWSXZHYrLl0wyzE4vsFdNru61rrgyN8pNDCwvTgFeo8",
        directory="static/gsheet",
        file_name="employee",
        file_extension="xlsx"
    )
    df = scrap_employee_xlsx(
        input_file_path="static/gsheet/employee.xlsx",
        sheet_name="Summary",
    )
    setup_employee_database(df)


if __name__ == "__main__":
    run_employee_data_pipeline()
